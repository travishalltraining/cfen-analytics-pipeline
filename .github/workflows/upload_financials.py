"""
CFEN Financial Data Upload Script
==================================
Reads a QuickBooks P&L Excel export and writes it to BigQuery.

Usage:
    python scripts/upload_financials.py "Financial_Reports_Review_-_Mar_2026.xlsx"

Requirements:
    pip install openpyxl google-cloud-bigquery

Environment variables required (same as sync_wodify_v2.py):
    GCP_PROJECT_ID
"""

import sys
import os
from datetime import datetime, timezone
from google.cloud import bigquery
import openpyxl

PROJECT_ID = os.environ["GCP_PROJECT_ID"]
DATASET_ID = "cfen_analytics"
TABLE_ID   = "monthly_financials"

BQ = bigquery.Client(project=PROJECT_ID)


def parse_pnl(filepath):
    """Parse a QuickBooks P&L Excel export and return list of monthly rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find P&L sheet
    sheet_name = next(
        (s for s in wb.sheetnames if 'p&l' in s.lower() or 'profit' in s.lower()),
        wb.sheetnames[0]
    )
    ws = wb[sheet_name]
    print(f"Reading sheet: '{sheet_name}'")

    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    # Find header row — contains month names like "Jan 2026"
    MONTH_ABBREVS = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        for j, cell in enumerate(row):
            if cell and any(str(cell).lower().startswith(m) for m in MONTH_ABBREVS):
                # Verify numeric data below
                if i + 2 < len(rows) and isinstance(rows[i+2][j], (int, float)):
                    header_idx = i
                    break
        if header_idx >= 0:
            break

    if header_idx < 0:
        raise ValueError("Could not find month columns in this file.")

    header = rows[header_idx]
    months = []
    total_col = -1
    for i, h in enumerate(header):
        if not h:
            continue
        s = str(h).strip()
        if s.lower() == 'total':
            total_col = i
        elif any(s.lower().startswith(m) for m in MONTH_ABBREVS):
            months.append({'name': s, 'col': i})

    print(f"Found months: {[m['name'] for m in months]}")

    def find(*labels):
        for label in labels:
            for row in rows:
                for ci in range(min(len(row), 2)):
                    if row[ci] is not None and str(row[ci]).strip().lower() == label.lower():
                        vals = {}
                        for m in months:
                            v = row[m['col']]
                            vals[m['name']] = float(v) if isinstance(v, (int, float)) else 0.0
                        if total_col >= 0:
                            tv = row[total_col]
                            vals['Total'] = float(tv) if isinstance(tv, (int, float)) else 0.0
                        else:
                            vals['Total'] = sum(vals[m['name']] for m in months)
                        return vals
        return None

    # Pull all key line items
    revenue    = find('Total for Income', 'Gross Profit', 'Total Income')
    expenses   = find('Total for Expenses', 'Total Expenses')
    net_income = find('Net Income', 'Net Profit')
    payroll    = find('Total for Payroll expenses', 'Total Payroll')
    rent       = find('Total for Rent', 'Rent')
    adv        = find('Total for Advertising & marketing', 'Advertising & marketing')
    membership = find('Membership')
    sgpt       = find('Semi Private Training')
    drop_in    = find('Drop In Sales')
    elements   = find('Elements Training')
    vending    = find('Vending Sales')
    mobile     = find('Mobile Private Training')
    apparel    = find('Apparel Sales')
    personal_tr= find('Personal Training')
    coaching   = find('Coaching Pay')
    admin_pay  = find('Administrative Pay', 'Admin Pay')
    supplies   = find('Total for Supplies', 'Supplies')
    software   = find('Total for Office expenses', 'Software & apps')
    insurance  = find('Insurance')
    utilities  = find('Total for Utilities', 'Utilities')
    gym_equip  = find('Total for Crossfit Expenses', 'Gym Equipment')
    other_exp  = find('Other Expenses')
    other_inc  = find('Interest earned', 'Other Income')

    if not revenue:
        raise ValueError("Could not find revenue data in this file.")

    synced_at = datetime.now(timezone.utc).isoformat()
    bq_rows = []

    for m in months:
        mn = m['name']
        # Parse month name and year
        parts = mn.split()
        month_name = parts[0] if parts else mn
        year_val = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 2026

        rev = revenue.get(mn, 0)
        exp = abs(expenses.get(mn, 0)) if expenses else 0
        net = net_income.get(mn, 0) if net_income else (rev - exp)

        row = {
            'report_month':            month_name,
            'report_year':             year_val,
            'membership':              membership.get(mn, 0) if membership else 0,
            'semi_private_training':   sgpt.get(mn, 0) if sgpt else 0,
            'elements_training':       elements.get(mn, 0) if elements else 0,
            'mobile_private_training': mobile.get(mn, 0) if mobile else 0,
            'drop_in_sales':           drop_in.get(mn, 0) if drop_in else 0,
            'vending_sales':           vending.get(mn, 0) if vending else 0,
            'apparel_sales':           apparel.get(mn, 0) if apparel else 0,
            'personal_training':       personal_tr.get(mn, 0) if personal_tr else 0,
            'other_income':            other_inc.get(mn, 0) if other_inc else 0,
            'payroll':                 abs(payroll.get(mn, 0)) if payroll else 0,
            'rent':                    abs(rent.get(mn, 0)) if rent else 0,
            'advertising_marketing':   abs(adv.get(mn, 0)) if adv else 0,
            'coaching_pay':            abs(coaching.get(mn, 0)) if coaching else 0,
            'admin_pay':               abs(admin_pay.get(mn, 0)) if admin_pay else 0,
            'gym_equipment':           abs(gym_equip.get(mn, 0)) if gym_equip else 0,
            'insurance':               abs(insurance.get(mn, 0)) if insurance else 0,
            'utilities':               abs(utilities.get(mn, 0)) if utilities else 0,
            'supplies':                abs(supplies.get(mn, 0)) if supplies else 0,
            'software_apps':           abs(software.get(mn, 0)) if software else 0,
            'other_expenses':          abs(other_exp.get(mn, 0)) if other_exp else 0,
            'total_revenue':           rev,
            'total_expenses':          exp,
            'net_income':              net,
            'synced_at':               synced_at,
        }
        bq_rows.append(row)
        print(f"  {mn}: Revenue=${rev:,.0f} | Expenses=${exp:,.0f} | Net=${net:,.0f}")

    return bq_rows


def upload_to_bigquery(rows):
    """Delete existing rows for these months and insert new ones."""
    table_ref = f"{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}"

    for row in rows:
        # Delete existing row for this month/year
        delete_sql = f"""
            DELETE FROM `{table_ref}`
            WHERE report_month = '{row['report_month']}'
              AND report_year = {row['report_year']}
        """
        BQ.query(delete_sql).result()
        print(f"Deleted existing data for {row['report_month']} {row['report_year']}")

    # Insert all rows
    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND
    )
    job = BQ.load_table_from_json(rows, table_ref, job_config=job_config)
    job.result()
    print(f"\n✅ Successfully uploaded {len(rows)} months to BigQuery!")


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/upload_financials.py <path_to_pnl.xlsx>")
        print("Example: python scripts/upload_financials.py 'Financial_Reports_Q1_2026.xlsx'")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    print(f"\n📊 Parsing P&L: {filepath}")
    rows = parse_pnl(filepath)

    print(f"\n📤 Uploading {len(rows)} months to BigQuery...")
    upload_to_bigquery(rows)

    print(f"\n🎉 Done! Data is now live in your dashboard.")


if __name__ == "__main__":
    main()
